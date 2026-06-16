from __future__ import annotations

import argparse
import os
from pathlib import Path

import pandas as pd

from .flatten import FlattenOptions, flatten_wiring_output
from .oracle import OracleConfig, fetch_query_dataframe


def _read_sql(path: Path) -> str:
    return path.read_text(encoding="utf-8")


def _default_output_path(input_path: Path, suffix: str) -> Path:
    return input_path.with_name(f"{input_path.stem}{suffix}{input_path.suffix}")

def _split_csv_cell(value) -> list[str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    text = str(value).strip()
    if not text:
        return []
    return [p.strip() for p in text.split(",") if p.strip()]

def _split_merged_cell(value) -> list[str]:
    """
    Split cells that may use our merge delimiter ('|') and/or comma lists.
    Example: '4, 5 | 6' -> ['4', '5', '6'].
    """
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    text = str(value).strip()
    if not text:
        return []
    parts: list[str] = []
    for chunk in text.split("|"):
        parts.extend(_split_csv_cell(chunk))
    return parts


def _explode_node0_terminals(df: pd.DataFrame) -> pd.DataFrame:
    """
    Expand output into multiple rows per instrument by splitting NODE0 terminals.

    This is a lightweight "loop drawing style" view: it does not attempt to
    thread each core through every intermediate termination; it only separates
    the instrument-side terminals (+/- etc.) into separate rows.
    """
    if "NODE0_TERMINALS" not in df.columns:
        return df
    rows: list[dict] = []
    for _, r in df.iterrows():
        terms = _split_csv_cell(r.get("NODE0_TERMINALS"))
        colors = _split_csv_cell(r.get("NODE0_COLOR_TERMINALS"))
        if not terms:
            rows.append(r.to_dict())
            continue

        for i, t in enumerate(terms):
            new_row = r.to_dict()
            new_row["NODE0_TERMINALS"] = t
            if colors:
                new_row["NODE0_COLOR_TERMINALS"] = colors[i] if i < len(colors) else colors[-1]
            rows.append(new_row)
    return pd.DataFrame.from_records(rows, columns=df.columns)

def _explode_all_nodes_by_index(df: pd.DataFrame, *, max_node: int | None = None) -> pd.DataFrame:
    """
    Expand output into multiple rows per instrument by splitting NODE{n} terminals/colors.

    This avoids repeating merged terminal lists across the multi-row view. It does
    not attempt a cartesian product; instead it creates up to N rows where N is
    the maximum terminal-count seen across NODE columns for that instrument row.
    """
    node_term_cols = [c for c in df.columns if c.startswith("NODE") and c.endswith("_TERMINALS")]
    if max_node is not None:
        filtered: list[str] = []
        for c in node_term_cols:
            # NODE0_TERMINALS, NODE1_TERMINALS, ...
            try:
                n = int(c[4 : c.find("_")])
            except Exception:
                continue
            if n <= max_node:
                filtered.append(c)
        node_term_cols = filtered
    if not node_term_cols:
        return df

    rows: list[dict] = []
    node_color_cols = [
        c.replace("_TERMINALS", "_COLOR_TERMINALS")
        for c in node_term_cols
        if c.replace("_TERMINALS", "_COLOR_TERMINALS") in df.columns
    ]
    for _, r in df.iterrows():
        per_node_terms: dict[str, list[str]] = {}
        per_node_colors: dict[str, list[str]] = {}
        per_node_strips: dict[str, list[str]] = {}
        max_len = 0
        for term_col in node_term_cols:
            node_prefix = term_col[: term_col.rfind("_TERMINALS")]  # e.g. NODE1
            color_col = f"{node_prefix}_COLOR_TERMINALS"
            strip_col = f"{node_prefix}_TERMINAL_STRIP"
            terms = _split_merged_cell(r.get(term_col))
            colors = _split_merged_cell(r.get(color_col)) if color_col in df.columns else []
            strips = _split_merged_cell(r.get(strip_col)) if strip_col in df.columns else []
            per_node_terms[node_prefix] = terms
            per_node_colors[node_prefix] = colors
            per_node_strips[node_prefix] = strips
            max_len = max(max_len, len(terms), len(colors))

        if max_len <= 1:
            rows.append(r.to_dict())
            continue

        for idx in range(max_len):
            new_row = r.to_dict()
            for node_prefix in per_node_terms.keys():
                term_col = f"{node_prefix}_TERMINALS"
                color_col = f"{node_prefix}_COLOR_TERMINALS"
                strip_col = f"{node_prefix}_TERMINAL_STRIP"
                terms = per_node_terms[node_prefix]
                colors = per_node_colors[node_prefix]
                strips = per_node_strips.get(node_prefix, [])
                new_row[term_col] = terms[idx] if idx < len(terms) else None
                if color_col in df.columns:
                    new_row[color_col] = colors[idx] if idx < len(colors) else None
                if strip_col in df.columns and strips:
                    if len(strips) > 1:
                        # Common case: one logical node panel but multiple strips.
                        # Keep "normal" core terminals on the first strip and push
                        # any extra (shield/drain) rows to the last strip.
                        chosen = strips[0] if idx < len(terms) else strips[-1]
                    else:
                        chosen = strips[0]
                    # Special-case: if this exploded row represents a shield/drain ("Metal"),
                    # show it on the last strip when multiple strips are present.
                    color_val = new_row.get(color_col)
                    if (
                        color_val is not None
                        and not (isinstance(color_val, float) and pd.isna(color_val))
                        and "metal" in str(color_val).lower()
                        and len(strips) > 1
                    ):
                        chosen = strips[-1]
                    new_row[strip_col] = chosen
            # Skip fully-empty exploded rows.
            has_any = False
            for c in node_term_cols:
                v = new_row.get(c)
                if v is not None and not (isinstance(v, float) and pd.isna(v)) and str(v).strip() not in ("", "nan", "None"):
                    has_any = True
                    break
            if not has_any:
                for c in node_color_cols:
                    v = new_row.get(c)
                    if v is not None and not (isinstance(v, float) and pd.isna(v)) and str(v).strip() not in ("", "nan", "None"):
                        has_any = True
                        break
            if has_any:
                rows.append(new_row)
    return pd.DataFrame.from_records(rows, columns=df.columns)


def _collapse_late_node_strips(df: pd.DataFrame, *, after_node: int) -> pd.DataFrame:
    """
    For nodes after `after_node`, collapse merged strip strings like "BAI1 | AI201"
    to the first strip ("BAI1") to keep the multi-row view focused on the connected strip.
    """
    out = df.copy()
    for col in out.columns:
        if not col.startswith("NODE"):
            continue
        if not (
            col.endswith("_TERMINAL_STRIP")
            or col.endswith("_TERMINALS")
            or col.endswith("_COLOR_TERMINALS")
        ):
            continue
        try:
            n = int(col[4 : col.find("_")])
        except Exception:
            continue
        if n <= after_node:
            continue
        series = out[col]
        if not pd.api.types.is_object_dtype(series.dtype) and not pd.api.types.is_string_dtype(series.dtype):
            continue
        out[col] = (
            series.astype("string")
            .str.split("|")
            .str[0]
            .astype("string")
            .str.strip()
            .where(series.notna(), other=series)
        )
    return out


def _style_xlsx(path: Path) -> None:
    if path.suffix.lower() != ".xlsx" or not path.exists():
        return
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception:
        return

    wb = load_workbook(path)
    ws = wb.active
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Light auto-fit (cap width) for readability.
    max_width = 45
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=min(ws.max_row, 200)), start=1):
        values = [c.value for c in col_cells[:50]]
        width = max((len(str(v)) for v in values if v is not None), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, width + 2), max_width)

    wb.save(path)


def cmd_flatten_excel(args: argparse.Namespace) -> int:
    input_path = Path(args.input)
    if not input_path.exists():
        raise SystemExit(f"Input Excel not found: {input_path}")

    df = pd.read_excel(input_path, sheet_name=args.sheet)
    key = ("LOOP_NAME", "INSTRUMENT_NAME") if args.key == "loop+instrument" else ("INSTRUMENT_NAME",)
    max_segments = _resolve_max_segments(df, key=key, value=args.max_segments, margin=args.auto_segments_margin)
    out_df = flatten_wiring_output(df, FlattenOptions(key=key, max_segments=max_segments, fixed_width=args.fixed_width))

    output_path = Path(args.output) if args.output else _default_output_path(input_path, "_flattened")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_df.to_excel(output_path, index=False)
    _style_xlsx(output_path)
    print(f"Wrote: {output_path}")
    return 0


def cmd_flatten_excel_multirows(args: argparse.Namespace) -> int:
    input_path = Path(args.input)
    if not input_path.exists():
        raise SystemExit(f"Input Excel not found: {input_path}")

    df = pd.read_excel(input_path, sheet_name=args.sheet)
    key = ("LOOP_NAME", "INSTRUMENT_NAME") if args.key == "loop+instrument" else ("INSTRUMENT_NAME",)
    max_segments = _resolve_max_segments(df, key=key, value=args.max_segments, margin=args.auto_segments_margin)
    out_df = flatten_wiring_output(df, FlattenOptions(key=key, max_segments=max_segments, fixed_width=args.fixed_width))
    # Expand all nodes by terminal index: one row per wire, common cells duplicated.
    out_df = _explode_all_nodes_by_index(out_df, max_node=None)

    output_path = (
        Path(args.output)
        if args.output
        else _default_output_path(input_path, "_flattened_multirows")
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_df.to_excel(output_path, index=False)
    _style_xlsx(output_path)
    print(f"Wrote: {output_path}")
    return 0


def cmd_flatten_db(args: argparse.Namespace) -> int:
    sql_path = Path(args.sql)
    if not sql_path.exists():
        raise SystemExit(f"SQL file not found: {sql_path}")

    user = args.user or os.getenv("ORACLE_USER")
    password = args.password or os.getenv("ORACLE_PASSWORD")
    dsn = args.dsn or os.getenv("ORACLE_DSN")
    if not user or not password or not dsn:
        raise SystemExit("Oracle credentials missing. Provide --user/--password/--dsn or env vars ORACLE_USER/ORACLE_PASSWORD/ORACLE_DSN.")

    unit_name_like = args.unit_name_like
    params = {"unit_name_like": unit_name_like}  # bind variable used by the provided SQL

    df = fetch_query_dataframe(
        config=OracleConfig(user=user, password=password, dsn=dsn),
        sql_text=_read_sql(sql_path),
        params=params,
        arraysize=args.arraysize,
    )

    key = ("LOOP_NAME", "INSTRUMENT_NAME") if args.key == "loop+instrument" else ("INSTRUMENT_NAME",)
    max_segments = _resolve_max_segments(df, key=key, value=args.max_segments, margin=args.auto_segments_margin)
    out_df = flatten_wiring_output(df, FlattenOptions(key=key, max_segments=max_segments, fixed_width=args.fixed_width))

    output_path = Path(args.output) if args.output else Path("wiring_retrieve_flattened.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    out_df.to_excel(output_path, index=False)
    _style_xlsx(output_path)
    print(f"Wrote: {output_path}")
    return 0


def _resolve_max_segments(df: pd.DataFrame, *, key: tuple[str, ...], value: str, margin: int) -> int:
    if value != "auto":
        try:
            resolved = int(value)
        except ValueError as exc:
            raise SystemExit(f"--max-segments must be an integer or 'auto' (got: {value!r})") from exc
        if resolved <= 0:
            raise SystemExit("--max-segments must be > 0 (or use 'auto').")
        return resolved

    # Upper-bound: number of input rows per instrument key (merges can only reduce hops).
    for col in key:
        if col not in df.columns:
            raise SystemExit(f"Missing key column in input: {col}")
    max_rows = int(df.groupby(list(key), dropna=False).size().max()) if len(df) else 0
    if max_rows <= 0:
        return 1
    return max(1, max_rows + int(margin))


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser(prog="loop-extractor", description="Flatten SPI wiring retrieve output to 1 row per instrument.")
    sub = p.add_subparsers(dest="cmd", required=True)

    px = sub.add_parser("flatten-excel", help="Flatten an exported Excel file.")
    px.add_argument("--input", required=True, help="Path to the wiring retrieve output .xlsx")
    px.add_argument("--sheet", default=0, help="Sheet name or index (default: 0)")
    px.add_argument("--output", help="Output .xlsx path (default: <input>_flattened.xlsx)")
    px.add_argument("--key", choices=["instrument", "loop+instrument"], default="instrument", help="Grouping key (default: instrument)")
    px.add_argument("--max-segments", default="10", help="Max segments per instrument, integer or 'auto' (default: 10)")
    px.add_argument("--auto-segments-margin", type=int, default=0, help="When --max-segments=auto, add this margin (default: 0)")
    px.add_argument("--fixed-width", action="store_true", help="Always output full max-segments columns (no trimming).")
    px.set_defaults(func=cmd_flatten_excel)

    pxr = sub.add_parser(
        "flatten-excel-multirows",
        help="Flatten and also split NODE0 terminals into multiple rows (loop-drawing style).",
    )
    pxr.add_argument("--input", required=True, help="Path to the wiring retrieve output .xlsx")
    pxr.add_argument("--sheet", default=0, help="Sheet name or index (default: 0)")
    pxr.add_argument("--output", help="Output .xlsx path (default: <input>_flattened_multirows.xlsx)")
    pxr.add_argument("--key", choices=["instrument", "loop+instrument"], default="instrument", help="Grouping key (default: instrument)")
    pxr.add_argument("--max-segments", default="10", help="Max segments per instrument, integer or 'auto' (default: 10)")
    pxr.add_argument("--auto-segments-margin", type=int, default=0, help="When --max-segments=auto, add this margin (default: 0)")
    pxr.add_argument("--fixed-width", action="store_true", help="Always output full max-segments columns (no trimming).")
    pxr.set_defaults(func=cmd_flatten_excel_multirows)

    pd_ = sub.add_parser("flatten-db", help="Run the provided SQL against Oracle (read-only) and flatten the results.")
    pd_.add_argument("--sql", default=str(Path("Retrieve") / "wiring_retrive.sql"), help="Path to the existing SQL file")
    pd_.add_argument("--dsn", help="Oracle DSN (or env ORACLE_DSN)")
    pd_.add_argument("--user", help="Oracle username (or env ORACLE_USER)")
    pd_.add_argument("--password", help="Oracle password (or env ORACLE_PASSWORD)")
    pd_.add_argument("--unit-name-like", default=None, help="Bind for :unit_name_like (example: 'UNIT%%')")
    pd_.add_argument("--arraysize", type=int, default=1000, help="Fetch arraysize (default: 1000)")
    pd_.add_argument("--output", default=None, help="Output .xlsx path (default: wiring_retrieve_flattened.xlsx)")
    pd_.add_argument("--key", choices=["instrument", "loop+instrument"], default="instrument", help="Grouping key (default: instrument)")
    pd_.add_argument("--max-segments", default="10", help="Max segments per instrument, integer or 'auto' (default: 10)")
    pd_.add_argument("--auto-segments-margin", type=int, default=0, help="When --max-segments=auto, add this margin (default: 0)")
    pd_.add_argument("--fixed-width", action="store_true", help="Always output full max-segments columns (no trimming).")
    pd_.set_defaults(func=cmd_flatten_db)

    return p


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
